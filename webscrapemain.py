import requests
from bs4 import BeautifulSoup

from openpyxl import Workbook
import time

book = Workbook()


# soup = BeautifulSoup(html_doc, 'html.parser')

url="https://www.bankswiftcode.org"

letter_urls = []
country_urls = []

# Make a GET request to fetch the raw HTML content
html_content = requests.get(url).text

# Parse the html content
soup = BeautifulSoup(html_content, "lxml")
print( type( soup ) )

uls = soup.find('ul', id="breadcrumbs")


xs = uls.find_all('a')

for x in xs:
    letter_urls.append( url + x.get('href') )
    # print( url + x.get('href') )

print( letter_urls )


for u in letter_urls:
    # print( link )
    html_content1 = requests.get( u ).text
    # print( html_content1 )
    soup1 = BeautifulSoup(html_content1, "lxml")

    # links = soup1.find('table', width="781")
    # links1 = links.find_all('td')
    links2 = soup1.find_all('a')

    for link in links2:
        if link.parent.name == 'td':
            # print( link["href"] )
            country_urls.append( url + link["href"] )

    # rows = table1.findAll('td')

    # for link in links2:
    #     print( link )


print( country_urls ) # print the parsed data of html

print( len(country_urls) )

del( country_urls[-7] )

for url in country_urls:

    # print(url)

    # print(index)

    html_content2 = requests.get( url ).text

    print( html_content2 )

    soup2 = BeautifulSoup(html_content2, "lxml")

    output = soup2.find('table', style="border: 5px #ccc solid;", width="774", cellpadding="10")

    country_element = soup2.find('p', id="heading")


    # print( country_element.text )

    if not output:
        output = soup2.find('table', id="t2")

    rows = [
    ('id','BankName', 'City', 'Branch', 'SwiftCode',	'Country', 'logourl', 'websiteurl', 'national identifier', 'bank routing scheme', 'bank routing address' )
    ]

    book.create_sheet(country_element.text)
    sheet = book.get_sheet_by_name(country_element.text)

    try:
        output1 = output.find_all('tr')    
    except AttributeError:
        continue

    if output1:
        for out in output1[1:]:
            # sp = BeautifulSoup(out, "lxml")
            tds = out.find_all('td')
            # print( tds )

            # for td in tds:
            try:
                record = (tds[0].text,tds[1].text, tds[2].text, tds[3].text, tds[4].text, country_element.text, None, None, None, None, None )
                rows.append( record )
                # print( td.text )
            except IndexError:
                continue
    
    for row in rows:
        sheet.append(row)
# sheet = book.active


# book.create_sheet("April")

# sheet['A1'] = 56
# sheet['A2'] = 43

# now = time.strftime("%x")
# sheet['A3'] = now

first_sheet = book.get_sheet_by_name('Sheet')
book.remove_sheet(first_sheet)

book.save("BankSwift.xlsx")

    

# print( output1 )

# print( type(output1) )